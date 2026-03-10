"""
Report Generator:
Creates a feasibility Excel report matching the exact format of
F-DEV-07 FEASIBILITY (Shri Ganesh Enterprises format)
Columns: Sr No | Description | Specification | Criticality | Proposed Machine |
         Inhouse/Outsource | Feasible | Reason Not Feasible | Deviation | Instrument |
         Inspection Inhouse | Frequency | Gauges Required | Remarks
"""
import openpyxl
from openpyxl.styles import (
    Font, Alignment, Border, Side, PatternFill, GradientFill
)
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import List, Dict, Any


HEADER_COLS = [
    "Sr No", "Description", "Specification", "Criticality\n(I/SC/CR)",
    "Proposed Machine", "Inhouse/\nOutsource", "Feasible\n(Yes/No)",
    "Reason For Not Feasible", "Deviation Required\nin Tolerance",
    "Measuring Type/\nInstruments", "Inhouse/\nOutsource",
    "Inspection\nFrequency", "Gauges Required\nfor Mass Production", "Remarks"
]

COL_WIDTHS = [8, 22, 18, 12, 18, 12, 10, 22, 22, 22, 12, 14, 20, 18]


def _thin_border():
    thin = Side(style="thin")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _header_fill():
    return PatternFill("solid", fgColor="1F4E79")


def _subheader_fill():
    return PatternFill("solid", fgColor="2E75B6")


def _alt_row_fill():
    return PatternFill("solid", fgColor="D9E1F2")


def generate_report(rfq_data: Dict, features: List[Dict], output_path: str) -> str:
    wb = openpyxl.Workbook()

    # ── Sheet 1: FEASIBILITY ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = "FEASIBILITY"

    # Company header
    ws.merge_cells("A1:N1")
    c = ws["A1"]
    c.value = "SHRI GANESH ENTERPRISES"
    c.font = Font(bold=True, size=14, color="FFFFFF")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = PatternFill("solid", fgColor="1F4E79")
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:K2")
    c = ws["A2"]
    c.value = "Product Feasibility Review"
    c.font = Font(bold=True, size=12, color="FFFFFF")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = PatternFill("solid", fgColor="2E75B6")
    ws.merge_cells("L2:N2")
    c2 = ws["L2"]
    c2.value = "DOC-F/DEV/07 | REV NO-01 1.06.2018"
    c2.font = Font(bold=True, size=8, color="FFFFFF")
    c2.alignment = Alignment(horizontal="center", vertical="center")
    c2.fill = PatternFill("solid", fgColor="1F4E79")
    ws.row_dimensions[2].height = 18

    # RFQ info rows
    info = [
        ("Part Name:", rfq_data.get("part_name", ""), "Customer Name:", rfq_data.get("customer_name", "")),
        ("Part No:", rfq_data.get("part_no", ""), "Contact Person:", ""),
        ("Drg Rev No/Date:", rfq_data.get("drg_rev", ""), "Supplier Team:", ""),
        ("Date:", datetime.now().strftime("%d.%m.%Y"), "Quantity:", str(rfq_data.get("quantity", ""))),
    ]
    for row_idx, (l1, v1, l2, v2) in enumerate(info, start=3):
        ws.merge_cells(f"A{row_idx}:B{row_idx}")
        ws[f"A{row_idx}"] = l1
        ws[f"A{row_idx}"].font = Font(bold=True, size=10)
        ws.merge_cells(f"C{row_idx}:F{row_idx}")
        ws[f"C{row_idx}"] = v1
        ws[f"C{row_idx}"].font = Font(size=10)
        ws.merge_cells(f"G{row_idx}:H{row_idx}")
        ws[f"G{row_idx}"] = l2
        ws[f"G{row_idx}"].font = Font(bold=True, size=10)
        ws.merge_cells(f"I{row_idx}:N{row_idx}")
        ws[f"I{row_idx}"] = v2
        ws[f"I{row_idx}"].font = Font(size=10)
        ws.row_dimensions[row_idx].height = 15

    # Column headers
    header_row = 7
    sub_row = 8
    ws.row_dimensions[header_row].height = 36
    ws.row_dimensions[sub_row].height = 36

    for col_idx, (header, width) in enumerate(zip(HEADER_COLS, COL_WIDTHS), start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = Font(bold=True, size=9, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = _header_fill()
        cell.border = _thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Manufacturing subheader spanning cols E-I
    ws.merge_cells(f"E{header_row}:I{header_row}")
    ws[f"E{header_row}"].value = "Manufacturing"
    ws[f"E{header_row}"].font = Font(bold=True, size=9, color="FFFFFF")
    ws[f"E{header_row}"].alignment = Alignment(horizontal="center", vertical="center")
    ws[f"E{header_row}"].fill = _subheader_fill()
    ws[f"E{header_row}"].border = _thin_border()

    # Measuring subheader spanning cols J-M
    ws.merge_cells(f"J{header_row}:M{header_row}")
    ws[f"J{header_row}"].value = "Measuring"
    ws[f"J{header_row}"].font = Font(bold=True, size=9, color="FFFFFF")
    ws[f"J{header_row}"].alignment = Alignment(horizontal="center", vertical="center")
    ws[f"J{header_row}"].fill = _subheader_fill()
    ws[f"J{header_row}"].border = _thin_border()

    # Data rows
    data_start = header_row + 1
    for i, feat in enumerate(sorted(features, key=lambda x: x.get("balloon_no", 0))):
        r = data_start + i
        fill = _alt_row_fill() if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        row_vals = [
            f"{feat.get('balloon_no', i+1)}.0",
            feat.get("description", ""),
            feat.get("specification", ""),
            feat.get("criticality", ""),
            feat.get("proposed_machine", ""),
            feat.get("inhouse_outsource", "Inhouse"),
            feat.get("feasible", "Yes"),
            feat.get("reason_not_feasible", ""),
            feat.get("deviation_required", ""),
            feat.get("measuring_instrument", ""),
            feat.get("inspection_inhouse", "Inhouse"),
            feat.get("inspection_frequency", ""),
            feat.get("gauge_required", ""),
            feat.get("remarks", ""),
        ]
        for col_idx, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=r, column=col_idx, value=val)
            cell.font = Font(size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.fill = fill
            cell.border = _thin_border()
            # Highlight infeasible rows
            if feat.get("feasible") == "No":
                cell.fill = PatternFill("solid", fgColor="FFD7D7")
        ws.row_dimensions[r].height = 18

    # Approval footer
    footer_row = data_start + len(features) + 2
    ws.merge_cells(f"A{footer_row}:D{footer_row}")
    ws[f"A{footer_row}"] = "Approved:                Not Approved:                Conditionally Approved:"
    ws[f"A{footer_row}"].font = Font(bold=True, size=9)

    sig_row = footer_row + 2
    ws.merge_cells(f"A{sig_row}:D{sig_row}")
    ws[f"A{sig_row}"] = "Sign:\nName:\nDate:"
    ws[f"A{sig_row}"].alignment = Alignment(wrap_text=True)
    ws.merge_cells(f"E{sig_row}:H{sig_row}")
    ws[f"E{sig_row}"] = "Sign:\nName:\nDate:"
    ws[f"E{sig_row}"].alignment = Alignment(wrap_text=True)
    ws.merge_cells(f"I{sig_row}:L{sig_row}")
    ws[f"I{sig_row}"] = "Sign:\nName:\nDate:"
    ws[f"I{sig_row}"].alignment = Alignment(wrap_text=True)
    ws.merge_cells(f"M{sig_row}:N{sig_row}")
    ws[f"M{sig_row}"] = "Approved By:"
    ws[f"M{sig_row}"].font = Font(bold=True)
    ws.row_dimensions[sig_row].height = 45

    # ── Sheet 2: BALLOONING DRG placeholder ──────────────────────────────────
    wb.create_sheet("BALLOONING DRG")
    bws = wb["BALLOONING DRG"]
    bws.merge_cells("A1:J1")
    bws["A1"] = "Balloon drawing"
    bws["A1"].font = Font(bold=True, size=14)
    bws.merge_cells("A3:J3")
    bws["A3"] = "See attached ballooned drawing image in the RFQ Agent system."
    bws["A3"].font = Font(size=11, italic=True)

    # ── Sheet 3: CEW-1 RMTC placeholder ──────────────────────────────────────
    wb.create_sheet("CEW-1 RMTC")
    rws = wb["CEW-1 RMTC"]
    rws["A1"] = "Raw Material Test Certificate"
    rws["A1"].font = Font(bold=True, size=14)
    rws["A2"] = "Attach supplier RMTC here"

    wb.save(output_path)
    print(f"[ReportGenerator] Excel saved: {output_path}")
    return output_path
