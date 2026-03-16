"""
AI Feasibility Report Agent
────────────────────────────
Reads ANY Excel template, understands its structure using Claude AI,
and fills it with extracted feature data + feasibility analysis.

Flow:
1. Parse template → extract structure (headers, merged cells, sample data)
2. Send structure + feature data to Claude → get fill instructions
3. Apply fill instructions to a copy of the template
4. Save filled report
"""

import os
import json
import copy
import re
import traceback
from typing import Dict, List, Any, Optional
from datetime import datetime

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import anthropic


CLAUDE_MODEL = "claude-sonnet-4-6"

# Sub-column order under merged group headers (well-known template convention)
MANUFACTURING_SUBCOLS = ["Proposed Machine", "Inhouse/Outsource", "Feasible (Yes/No)", "Reason if Not Feasible", "Deviation Required"]
MEASURING_SUBCOLS = ["Measuring Instrument", "Inhouse/Outsource", "Inspection Frequency", "Gauge Required"]


# ═══════════════════════════════════════════════════════════════════════════
# Step 1: Parse template structure
# ═══════════════════════════════════════════════════════════════════════════

def parse_template_structure(template_path: str) -> Dict[str, Any]:
    """
    Read an Excel template and extract its complete structure:
    - Sheet names
    - Per-sheet: merged cells, cell values, column widths, row heights
    - Identify header rows, data rows, repeating patterns
    """
    wb = openpyxl.load_workbook(template_path)
    structure = {"sheets": {}, "file": os.path.basename(template_path)}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_info = {
            "max_row": ws.max_row,
            "max_col": ws.max_column,
            "merged_cells": [str(m) for m in ws.merged_cells.ranges],
            "cells": {},
            "column_headers": [],
            "data_start_row": None,
            "sample_data_rows": [],
        }

        # Read all cells with values
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                if cell.value is not None:
                    coord = cell.coordinate
                    cell_info = {
                        "value": str(cell.value),
                        "row": cell.row,
                        "col": cell.column,
                        "col_letter": get_column_letter(cell.column),
                        "bold": cell.font.bold if cell.font else False,
                        "merged": any(coord in str(m) for m in ws.merged_cells.ranges),
                    }
                    sheet_info["cells"][coord] = cell_info

        # Detect header row pattern (row with most bold cells or cells with column-like headers)
        row_cell_counts = {}
        for coord, info in sheet_info["cells"].items():
            r = info["row"]
            if r not in row_cell_counts:
                row_cell_counts[r] = {"total": 0, "bold": 0, "values": []}
            row_cell_counts[r]["total"] += 1
            if info["bold"]:
                row_cell_counts[r]["bold"] += 1
            row_cell_counts[r]["values"].append(info["value"])

        # Find the row that looks like a data table header
        # (most columns filled, bold text, keywords like Sr No, Description, etc.)
        header_keywords = [
            "sr", "no", "description", "specification", "spec", "dimension",
            "tolerance", "machine", "process", "instrument", "remarks",
            "criticality", "feasible", "measuring", "manufacturing",
            "inhouse", "outsource", "frequency", "gauge",
        ]
        best_header_row = None
        best_score = 0
        for r, info in row_cell_counts.items():
            score = info["total"] + info["bold"] * 2
            for val in info["values"]:
                val_lower = val.lower()
                for kw in header_keywords:
                    if kw in val_lower:
                        score += 3
            if score > best_score and info["total"] >= 3:
                best_score = score
                best_header_row = r

        if best_header_row:
            # Extract column headers from that row
            headers = []
            for coord, info in sheet_info["cells"].items():
                if info["row"] == best_header_row:
                    headers.append({
                        "col": info["col"],
                        "col_letter": info["col_letter"],
                        "header": info["value"],
                    })
            headers.sort(key=lambda x: x["col"])
            sheet_info["column_headers"] = headers
            sheet_info["header_row"] = best_header_row

            # Find data start row (first row after header with data)
            for r in range(best_header_row + 1, ws.max_row + 1):
                has_data = False
                for coord, info in sheet_info["cells"].items():
                    if info["row"] == r:
                        has_data = True
                        break
                if has_data:
                    sheet_info["data_start_row"] = r
                    break

            # Grab sample data rows (up to 3)
            if sheet_info["data_start_row"]:
                for r in range(sheet_info["data_start_row"],
                               min(sheet_info["data_start_row"] + 3, ws.max_row + 1)):
                    row_data = {}
                    for coord, info in sheet_info["cells"].items():
                        if info["row"] == r:
                            row_data[info["col_letter"]] = info["value"]
                    if row_data:
                        sheet_info["sample_data_rows"].append({"row": r, "data": row_data})

        structure["sheets"][sheet_name] = sheet_info

    wb.close()
    return structure


# ═══════════════════════════════════════════════════════════════════════════
# Step 2: Build context for Claude
# ═══════════════════════════════════════════════════════════════════════════

def build_template_description(structure: Dict) -> str:
    """Convert parsed structure into a readable description for Claude."""
    lines = [f"Excel Template: {structure['file']}", ""]

    for sheet_name, sheet in structure["sheets"].items():
        lines.append(f"═══ Sheet: '{sheet_name}' (rows={sheet['max_row']}, cols={sheet['max_col']}) ═══")

        # Show all cells with values
        lines.append("All cells with content:")
        sorted_cells = sorted(sheet["cells"].items(),
                              key=lambda x: (x[1]["row"], x[1]["col"]))
        current_row = None
        for coord, info in sorted_cells:
            if info["row"] != current_row:
                current_row = info["row"]
                lines.append(f"  Row {current_row}:")
            bold_tag = " [BOLD]" if info["bold"] else ""
            merged_tag = " [MERGED]" if info["merged"] else ""
            lines.append(f"    {coord} = \"{info['value']}\"{bold_tag}{merged_tag}")

        # Show detected headers
        if sheet["column_headers"]:
            lines.append(f"\nDetected table header row: {sheet.get('header_row', '?')}")
            lines.append("Columns:")
            for h in sheet["column_headers"]:
                lines.append(f"  Column {h['col_letter']}: \"{h['header']}\"")

        # Show sample data
        if sheet["sample_data_rows"]:
            lines.append(f"\nSample data rows (starting row {sheet['data_start_row']}):")
            for sample in sheet["sample_data_rows"]:
                row_str = ", ".join(f"{k}={v}" for k, v in sorted(sample["data"].items()))
                lines.append(f"  Row {sample['row']}: {row_str}")

        # Merged cells
        if sheet["merged_cells"]:
            lines.append(f"\nMerged cell ranges: {', '.join(sheet['merged_cells'])}")

        lines.append("")

    return "\n".join(lines)


# ═══════════════════════════════════════════════════════════════════════════
# Step 3: Claude AI fills the template
# ═══════════════════════════════════════════════════════════════════════════

def generate_fill_instructions(
    template_desc: str,
    rfq_data: Dict,
    features: List[Dict],
    manufacturing_metadata: Optional[Dict] = None,
) -> Dict:
    """
    Send template structure + feature data to Claude.
    Claude returns JSON instructions for filling each cell.
    """
    client = anthropic.Anthropic()

    feature_summary = json.dumps(features, indent=2, default=str)
    rfq_summary = json.dumps(rfq_data, indent=2, default=str)
    meta_summary = json.dumps(manufacturing_metadata, indent=2, default=str) if manufacturing_metadata else "None"

    prompt = f"""You are a manufacturing feasibility report expert. Fill an Excel template with drawing feature data.

TEMPLATE STRUCTURE:
{template_desc}

RFQ DATA:
{rfq_summary}

EXTRACTED FEATURES ({len(features)} total — you MUST include every single one):
{feature_summary}

MANUFACTURING METADATA:
{meta_summary}

═══ CRITICAL RULES ═══

1. EVERY feature MUST appear as a data_row. You have {len(features)} features → you MUST return exactly {len(features)} data_rows. NEVER skip, merge, or truncate features.

2. If the template has fewer empty rows than features (e.g. 14 rows for 17 features), CONTINUE with sequential row numbers beyond the template. The code inserts rows automatically.

3. ZERO ABBREVIATIONS — copy values EXACTLY from feature data:
   - measuring_instrument: use the EXACT string (e.g. "DIGITAL HEIGHT GAUGE" not "DHG", "DIGITAL VERNIER CALIPER" not "DVC", "Micrometer" not "Mic")
   - inhouse_outsource / inspection_inhouse: write "Inhouse" or "Outsource" in full (NEVER "IN" or "Out" or "I/H")
   - proposed_machine: use EXACT string from data (e.g. "CNC LATHE" not "CNC")
   - All other fields: copy verbatim from the feature data, do not paraphrase or shorten

4. Sr No column: plain integers "1", "2", "3" — never "1.0" or "01"

5. Date: {datetime.now().strftime("%d.%m.%Y")}

═══ COLUMN MAPPING ═══

The template uses merged group headers. Here is the sub-column layout:

Manufacturing group (5 sub-columns in order):
  Col E = proposed_machine | Col F = inhouse_outsource | Col G = feasible | Col H = reason_not_feasible | Col I = deviation_required

Measuring group (4 sub-columns in order):
  Col J = measuring_instrument | Col K = inspection_inhouse | Col L = inspection_frequency | Col M = gauge_required

Individual columns:
  A = balloon_no (integer) | B = description | C = specification | D = criticality | N = remarks

Header info cells (match label → value cell to the right):
  Part Name → part_name | Part No → part_no | Customer Name → customer_name | Date → today | Quantity → quantity | Drg Rev → drg_rev

═══ RESPONSE FORMAT ═══

Return ONLY valid JSON:
{{
  "header_fills": [
    {{"cell": "C3", "value": "..."}},
    {{"cell": "I3", "value": "..."}}
  ],
  "data_rows": [
    {{
      "row_number": 8,
      "cells": {{
        "A": "1",
        "B": "Total Length",
        "C": "87 ±0.5",
        "D": "SC",
        "E": "CNC LATHE",
        "F": "Inhouse",
        "G": "Yes",
        "H": "",
        "I": "",
        "J": "DIGITAL HEIGHT GAUGE",
        "K": "Inhouse",
        "L": "1/Hr",
        "M": "",
        "N": ""
      }}
    }}
  ],
  "data_start_row": 8,
  "clear_sample_rows": true
}}

FINAL CHECK before responding:
- Count your data_rows. Is it exactly {len(features)}? If not, you missed features — go back and add them.
- Did you use ANY abbreviations? Search for "DHG", "DVC", "IN", "Out" — replace with full words.
"""

    print(f"[FeasibilityAgent] Sending to Claude ({len(features)} features, template: {template_desc[:60]}...)")

    response = client.messages.create(
        model=CLAUDE_MODEL,
        max_tokens=16000,
        messages=[{"role": "user", "content": prompt}],
    )

    text = response.content[0].text.strip()

    # Extract JSON from response
    json_match = re.search(r'\{[\s\S]*\}', text)
    if json_match:
        instructions = json.loads(json_match.group())
        print(f"[FeasibilityAgent] Claude returned {len(instructions.get('data_rows', []))} data rows, "
              f"{len(instructions.get('header_fills', []))} header fills")
        return instructions
    else:
        raise ValueError(f"Claude did not return valid JSON. Response: {text[:500]}")


# ═══════════════════════════════════════════════════════════════════════════
# Step 4: Apply fill instructions to template copy
# ═══════════════════════════════════════════════════════════════════════════

def apply_fill_instructions(template_path: str, instructions: Dict, output_path: str, ballooned_image_path: str = None) -> str:
    """
    Copy the template and fill cells based on Claude's instructions.
    Preserves all formatting, merged cells, styles from original.
    """
    wb = openpyxl.load_workbook(template_path)

    # Strip Export Summary sheet added by Numbers app on export
    for name in list(wb.sheetnames):
        if "export summary" in name.lower():
            del wb[name]

    # Find the main feasibility sheet (first sheet or one with "feasib" in name)
    target_sheet = wb.sheetnames[0]
    for name in wb.sheetnames:
        if "feasib" in name.lower():
            target_sheet = name
            break
    ws = wb[target_sheet]

    # Apply header fills
    for fill in instructions.get("header_fills", []):
        cell_ref = fill["cell"]
        value = fill["value"]
        if value:
            target_cell = _get_writable_cell(ws, cell_ref)
            target_cell.value = value
            print(f"[FeasibilityAgent] Header: {cell_ref} = {value}")

    data_start = instructions.get("data_start_row", 8)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # ── Detect footer row (first row at/after data_start with approval/signature text) ──
    footer_kws = ["approved:", "not approved", "conditionally approved", "sign:", "name:", "sign:\n"]
    footer_row = None
    for r in range(data_start, ws.max_row + 1):
        for c in range(1, min(ws.max_column + 1, 6)):
            val = ws.cell(row=r, column=c).value
            if val and isinstance(val, str) and any(kw in val.lower() for kw in footer_kws):
                footer_row = r
                break
        if footer_row:
            break

    # ── Insert extra rows if features overflow the template's data zone ──
    n_features = len(instructions.get("data_rows", []))
    if footer_row:
        available = footer_row - data_start
        if n_features > available:
            extra = n_features - available
            # Snapshot the last template data row's styles before insertion
            src_row_idx = footer_row - 1
            src_styles = []
            for c in range(1, ws.max_column + 1):
                src = ws.cell(row=src_row_idx, column=c)
                src_styles.append({
                    "font": copy.copy(src.font),
                    "fill": copy.copy(src.fill),
                    "alignment": copy.copy(src.alignment),
                    "number_format": src.number_format,
                })
            ws.insert_rows(footer_row, extra)
            for offset in range(extra):
                new_r = footer_row + offset
                ws.row_dimensions[new_r].height = 18
                for c, style in enumerate(src_styles, start=1):
                    cell = ws.cell(row=new_r, column=c)
                    try:
                        cell.font = style["font"]
                        cell.fill = style["fill"]
                        cell.alignment = style["alignment"]
                        cell.number_format = style["number_format"]
                        cell.border = thin_border
                    except AttributeError:
                        pass  # MergedCell
            print(f"[FeasibilityAgent] Inserted {extra} rows before footer (row {footer_row})")

    # ── Clear existing sample data rows (data zone only, stop before footer) ──
    clear_limit = (footer_row if footer_row else ws.max_row + 1)
    if instructions.get("clear_sample_rows", False):
        merges_to_remove = [str(m) for m in ws.merged_cells.ranges if m.min_row >= data_start and m.max_row < clear_limit]
        for m in merges_to_remove:
            try:
                ws.unmerge_cells(m)
            except Exception:
                pass
        for row in range(data_start, clear_limit):
            for col in range(1, ws.max_column + 1):
                try:
                    ws.cell(row=row, column=col).value = None
                except AttributeError:
                    pass  # MergedCell

    # Deduplicate by Sr No (column A), normalise "1.0" → "1"
    seen_sr = set()
    deduped = []
    for row_data in instructions.get("data_rows", []):
        sr_raw = str(row_data.get("cells", {}).get("A", "")).strip()
        try:
            key = str(int(float(sr_raw))) if sr_raw else str(row_data.get("row_number", ""))
        except ValueError:
            key = sr_raw or str(row_data.get("row_number", ""))
        if key not in seen_sr:
            seen_sr.add(key)
            deduped.append(row_data)
    instructions["data_rows"] = deduped

    # Renumber rows sequentially from data_start (ignore Claude's row_number — it may
    # not account for inserted overflow rows or may skip numbers)
    for i, row_data in enumerate(instructions.get("data_rows", [])):
        row_data["row_number"] = data_start + i

    # Alternating row fills for readability
    alt_fill = PatternFill("solid", fgColor="D9E1F2")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    red_fill = PatternFill("solid", fgColor="FFD7D7")

    for i, row_data in enumerate(instructions.get("data_rows", [])):
        row_num = row_data["row_number"]
        cells = row_data.get("cells", {})
        is_infeasible = str(cells.get("G", "")).strip().lower() == "no"
        row_fill = red_fill if is_infeasible else (alt_fill if i % 2 == 0 else white_fill)

        for col_letter, value in cells.items():
            try:
                col_idx = column_index_from_string(col_letter)
                cell = ws.cell(row=row_num, column=col_idx)
                try:
                    # Normalise Sr No to plain int
                    if col_letter == "A" and value:
                        try:
                            value = str(int(float(value)))
                        except (ValueError, TypeError):
                            pass
                    cell.value = value if value else ""
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.font = Font(name="Arial", size=9)
                    cell.fill = row_fill
                except AttributeError:
                    pass  # MergedCell
            except Exception as e:
                print(f"[FeasibilityAgent] Warning: Could not fill {col_letter}{row_num}: {e}")

    print(f"[FeasibilityAgent] Wrote {len(instructions.get('data_rows', []))} feature rows")

    # Handle BALLOONING DRG sheet - insert image if available
    if ballooned_image_path and os.path.exists(ballooned_image_path):
        balloon_sheet = None
        for name in wb.sheetnames:
            if "balloon" in name.lower() or "drg" in name.lower():
                balloon_sheet = name
                break
        if balloon_sheet:
            bws = wb[balloon_sheet]
            try:
                from openpyxl.drawing.image import Image as XlImage
                img = XlImage(ballooned_image_path)
                img.width = 800
                img.height = int(800 * img.height / img.width) if img.width > 0 else 600
                bws.add_image(img, "A3")
                print(f"[FeasibilityAgent] Inserted ballooned image into '{balloon_sheet}' sheet")
            except Exception as e:
                print(f"[FeasibilityAgent] Could not insert image: {e}")

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    wb.close()
    print(f"[FeasibilityAgent] Report saved: {output_path}")
    return output_path


def _get_writable_cell(ws, cell_ref: str):
    """Get the actual writable cell, handling merged ranges."""
    cell = ws[cell_ref]
    # If the cell is in a merged range, write to the top-left cell of that range
    for merged_range in ws.merged_cells.ranges:
        if cell_ref in merged_range:
            return ws.cell(row=merged_range.min_row, column=merged_range.min_col)
    return cell


# ═══════════════════════════════════════════════════════════════════════════
# Main entry point
# ═══════════════════════════════════════════════════════════════════════════

def generate_feasibility_report(
    template_path: str,
    rfq_data: Dict,
    features: List[Dict],
    output_path: str,
    manufacturing_metadata: Optional[Dict] = None,
    ballooned_image_path: str = None,
) -> str:
    """
    Main function: reads any Excel template, uses Claude AI to understand
    its structure, maps feature data, and fills the report.

    Args:
        template_path: path to customer's Excel template
        rfq_data: dict with part_name, part_no, customer_name, drg_rev, quantity
        features: list of feature dicts from feasibility engine
        output_path: where to save the filled report
        manufacturing_metadata: optional metadata dict
        ballooned_image_path: optional path to ballooned drawing image

    Returns:
        output_path of the generated report
    """
    print(f"[FeasibilityAgent] ═══ Starting Report Generation ═══")
    print(f"[FeasibilityAgent] Template: {template_path}")
    print(f"[FeasibilityAgent] Features: {len(features)}")

    # Step 1: Parse template
    print(f"[FeasibilityAgent] Step 1: Parsing template structure...")
    structure = parse_template_structure(template_path)

    # Step 2: Build description
    template_desc = build_template_description(structure)
    print(f"[FeasibilityAgent] Template description: {len(template_desc)} chars")

    # Step 3: Get fill instructions from Claude
    print(f"[FeasibilityAgent] Step 2: Getting fill instructions from Claude AI...")
    instructions = generate_fill_instructions(
        template_desc, rfq_data, features, manufacturing_metadata
    )

    # Step 4: Apply to template
    print(f"[FeasibilityAgent] Step 3: Applying fill instructions to template...")
    result = apply_fill_instructions(template_path, instructions, output_path, ballooned_image_path)

    print(f"[FeasibilityAgent] ═══ Report Generation Complete ═══")
    return result
