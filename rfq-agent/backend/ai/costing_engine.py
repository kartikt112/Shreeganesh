"""
Cost Break-Up Sheet Engine for SGE CNC Manufacturing.

Generates per-piece cost breakdown:
  Section A: Raw Material Cost
  Section B: Process Cost (operations × machine rates)
  Section C: Overheads (configurable percentages)
  Section D: Tooling
  Total = A + B + C + D

Formulas match the template: 26_408_KA_Fork_Assembly_childpart.xlsx
"""
import math
import json
import os
import copy
import re
from typing import List, Dict, Any, Optional, Tuple

import anthropic
import openpyxl
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# 1.  RAW MATERIAL CALCULATIONS
# ---------------------------------------------------------------------------

def calculate_raw_material(
    rod_dia_mm: float,
    rod_length_mm: float,
    net_weight_kg: float,
    density_g_cm3: float,
    mat_rate_per_kg: float,
    zbc: float = 30.0,
    yield_pct: float = 85.0,
    bop_cost: float = 0.0,
    rod_shape: str = "ROD",
) -> Dict[str, Any]:
    """
    Section A: Raw Material Cost.
    Formulas from template:
      Gross Wt = C15*C15*D15*7.86*0.786/1000000   (for rod)
      Scrap = Gross - Net
      Part RM Rate = (Gross*MatRate) - (Scrap*ZBC)*0.85
    """
    # Gross weight calculation (template formula: Dia*Dia*Length*density*0.786/1e6)
    # This equals: π/4 * (Dia_mm)² * Length_mm * density / 1e6  (since π/4 ≈ 0.7854)
    gross_wt = rod_dia_mm * rod_dia_mm * rod_length_mm * density_g_cm3 * 0.786 / 1_000_000
    scrap = gross_wt - net_weight_kg

    # Part RM Rate formula from template: =((H14*H17)-(H16*E17)*0.85)
    # H14=gross_wt, H17=mat_rate, H16=scrap, E17=zbc
    part_rm_rate = (gross_wt * mat_rate_per_kg) - (scrap * zbc * 0.85)

    return {
        "rod_shape": rod_shape,
        "rod_dia_mm": rod_dia_mm,
        "rod_length_mm": rod_length_mm,
        "density_g_cm3": density_g_cm3,
        "gross_weight_kg": round(gross_wt, 6),
        "net_weight_kg": net_weight_kg,
        "scrap_kg": round(scrap, 6),
        "zbc": zbc,
        "yield_pct": yield_pct,
        "mat_rate_per_kg": mat_rate_per_kg,
        "part_rm_rate": round(part_rm_rate, 4),
        "bop_cost": bop_cost,
    }


# ---------------------------------------------------------------------------
# 2.  PROCESS COST CALCULATIONS
# ---------------------------------------------------------------------------

def calculate_process_cost(
    operations: List[Dict[str, Any]],
    plating_rate_per_kg: float = 0.0,
    net_weight_kg: float = 0.0,
    annealing_rate_per_sqin: float = 0.0,
    surface_area_sqin: float = 0.0,
) -> Dict[str, Any]:
    """
    Section B: Process Cost.
    Each operation: Cost = Rate / Strokes
    Plus plating and annealing if applicable.
    """
    total_process = 0.0
    ops_with_cost = []
    for op in operations:
        strokes = op.get("strokes_per_hr", 1)
        rate = op.get("rate_per_hr", 0)
        cost = rate / strokes if strokes > 0 else 0
        ops_with_cost.append({
            **op,
            "cost": round(cost, 4),
        })
        total_process += cost

    plating_cost = net_weight_kg * plating_rate_per_kg
    annealing_cost = surface_area_sqin * annealing_rate_per_sqin

    return {
        "operations": ops_with_cost,
        "total_process_cost": round(total_process, 4),
        "plating_cost": round(plating_cost, 4),
        "plating_rate_per_kg": plating_rate_per_kg,
        "annealing_cost": round(annealing_cost, 4),
        "annealing_rate_per_sqin": annealing_rate_per_sqin,
        "surface_area_sqin": surface_area_sqin,
    }


# ---------------------------------------------------------------------------
# 3.  OVERHEAD CALCULATIONS
# ---------------------------------------------------------------------------

def calculate_overheads(
    rm_cost: float,
    bop_cost: float,
    process_cost: float,
    plating_cost: float,
    annealing_cost: float,
    net_weight_kg: float,
    config: Dict[str, float],
) -> Dict[str, Any]:
    """
    Section C: Overheads.
    Formulas from template:
      Subtotal(G33) = RM + BOP + Process + Plating + Annealing
      Rejection = rejection_pct * G33
      ICC = icc_pct * (RM + BOP)
      Overheads = overheads_pct * G33
      Profit = profit_pct * RM
      Packing = packing_pct * G33
      Freight = freight_per_kg * net_weight
      Inspection = inspection_pct * process_cost
    """
    subtotal_g33 = rm_cost + bop_cost + process_cost + plating_cost + annealing_cost

    rejection = config.get("rejection_pct", 0.02) * subtotal_g33
    icc = config.get("icc_pct", 0.015) * (rm_cost + bop_cost)
    overheads = config.get("overheads_pct", 0.065) * subtotal_g33
    profit = config.get("profit_pct", 0.10) * rm_cost
    packing = config.get("packing_pct", 0.015) * subtotal_g33
    freight = config.get("freight_per_kg", 8.0) * net_weight_kg
    inspection = config.get("inspection_pct", 0.02) * process_cost

    overhead_total = rejection + icc + overheads + profit + packing + freight + inspection

    return {
        "subtotal_rm_process": round(subtotal_g33, 4),
        "rejection": round(rejection, 4),
        "rejection_pct": config.get("rejection_pct", 0.02),
        "icc": round(icc, 4),
        "icc_pct": config.get("icc_pct", 0.015),
        "overheads": round(overheads, 4),
        "overheads_pct": config.get("overheads_pct", 0.065),
        "profit": round(profit, 4),
        "profit_pct": config.get("profit_pct", 0.10),
        "packing": round(packing, 4),
        "packing_pct": config.get("packing_pct", 0.015),
        "freight": round(freight, 4),
        "freight_per_kg": config.get("freight_per_kg", 8.0),
        "inspection": round(inspection, 4),
        "inspection_pct": config.get("inspection_pct", 0.02),
        "overhead_total": round(overhead_total, 4),
    }


# ---------------------------------------------------------------------------
# 4.  TOTAL COST
# ---------------------------------------------------------------------------

def calculate_total(
    rm: Dict, process: Dict, overheads: Dict,
    tooling: Dict[str, float] = None,
) -> Dict[str, Any]:
    """Total Part Cost = A + B + C + D."""
    if tooling is None:
        tooling = {"hydraulic_fixture": 0, "gauges": 0, "investment_casting_tool": 0}

    part_cost_abc = overheads["subtotal_rm_process"] + overheads["overhead_total"]
    tool_total = sum(tooling.values())
    # Template: G50 = F50*G44  where F50 = total_tool_cost...
    # Actually G50 = 0 if tools=0. Tool cost is amortized per piece if quantity given.
    total = part_cost_abc + tool_total

    return {
        "raw_material": rm,
        "process": process,
        "overheads": overheads,
        "tooling": tooling,
        "part_cost_abc": round(part_cost_abc, 4),
        "tool_cost": round(tool_total, 4),
        "total_part_cost": round(total, 4),
    }


# ---------------------------------------------------------------------------
# 5.  AI OPERATION ESTIMATOR
# ---------------------------------------------------------------------------

def estimate_operations_from_features(
    features: List[Dict[str, Any]],
    metadata: Dict[str, Any],
    machine_rates: Dict[str, float],
) -> List[Dict[str, Any]]:
    """
    Use Claude AI to analyze drawing features and suggest manufacturing
    operations with cycle times. Falls back to rule-based if AI fails.
    """
    try:
        return _ai_estimate_operations(features, metadata, machine_rates)
    except Exception as e:
        print(f"[Costing] AI estimation failed ({e}), falling back to rules")
        return _rule_based_estimate(features, metadata, machine_rates)


def _ai_estimate_operations(
    features: List[Dict[str, Any]],
    metadata: Dict[str, Any],
    machine_rates: Dict[str, float],
) -> List[Dict[str, Any]]:
    """
    Call Claude to analyze features and generate a manufacturing process plan
    with operation sequence and realistic cycle time estimates.
    """
    client = anthropic.Anthropic()

    # Build feature summary for the prompt
    feature_lines = []
    for f in features:
        spec = f.get("specification") or f.get("spec") or ""
        ftype = f.get("feature_type") or f.get("type") or ""
        if spec:
            feature_lines.append(f"  #{f.get('balloon_no', '?')}: {ftype} — {spec}")

    feature_text = "\n".join(feature_lines) if feature_lines else "No features extracted"

    # Part info
    envelope = metadata.get("part_envelope", {})
    material_info = metadata.get("material", {})
    if isinstance(material_info, dict):
        mat_grade = material_info.get("grade", "Unknown")
    else:
        mat_grade = str(material_info)

    part_info = f"""Part: {metadata.get('part_name', 'Unknown')}
Drawing: {metadata.get('drawing_number', 'Unknown')}
Material: {mat_grade}
Max OD: {envelope.get('max_od_mm', 'Unknown')} mm
Total Length: {envelope.get('total_length_mm', 'Unknown')} mm
Is Hollow: {envelope.get('is_hollow', False)}
General Tolerance: {metadata.get('general_tolerance_standard', 'Unknown')}
Surface Protection: {metadata.get('surface_protection', {}).get('method', 'None') if isinstance(metadata.get('surface_protection'), dict) else 'None'}"""

    # Available machines and rates
    machine_lines = [f"  {name}: Rs {rate}/hr" for name, rate in machine_rates.items()]
    machines_text = "\n".join(machine_lines)

    prompt = f"""You are a CNC manufacturing process planner for a precision job shop (Shri Ganesh Enterprises, Aurangabad, India).

Given the part details and drawing features below, plan the manufacturing operation sequence.

{part_info}

Drawing Features:
{feature_text}

Available Machines & Rates:
{machines_text}

TASK: Determine the manufacturing operations needed to produce this part from bar stock.
For each operation, estimate realistic strokes/hour (parts per hour) based on the part complexity.

RULES:
- First operation is always Parting (cutting bar stock to length)
- For small parts (OD ≤ 25mm), use Traub for parting. For larger, use CNC Cutting.
- Group CNC turning operations into setups (1st Setup = chuck side, 2nd Setup = flip side)
- Consider: Does this part need drilling, tapping, threading, grinding, milling, plating?
- Strokes/hr = how many parts per hour for that operation. Consider:
  - Simple parting: 60-120/hr
  - CNC turning with multiple features: 15-30/hr per setup
  - Drilling/tapping: 30-60/hr
  - Thread rolling: 100-200/hr
  - Centerless grinding: 40-80/hr
  - VMC milling: 20-40/hr
- Choose machines from the available list above.
- Include plating/coating only if surface protection is specified.

Return ONLY a JSON array, no other text. Example:
```json
[
  {{"sno": 1, "process": "Parting", "machine": "Traub", "strokes_per_hr": 85}},
  {{"sno": 2, "process": "CNC 1st Setup", "machine": "CNC", "strokes_per_hr": 25}},
  {{"sno": 3, "process": "Drilling & Tapping", "machine": "VMC/Turret Lathe", "strokes_per_hr": 50}}
]
```"""

    print("[Costing] Calling Claude for operation estimation...")
    response = client.messages.create(
        model="claude-opus-4-6",
        max_tokens=1024,
        messages=[{"role": "user", "content": prompt}],
    )

    # Parse response
    text = response.content[0].text.strip()
    print(f"[Costing] Claude response: {text[:200]}...")

    # Extract JSON from response (handle markdown code blocks)
    json_match = re.search(r'```(?:json)?\s*(\[.*?\])\s*```', text, re.DOTALL)
    if json_match:
        ops_json = json.loads(json_match.group(1))
    elif text.startswith("["):
        ops_json = json.loads(text)
    else:
        raise ValueError(f"Could not parse operations JSON from response: {text[:100]}")

    # Attach machine rates
    operations = []
    for op in ops_json:
        rate = _find_rate(op.get("machine", ""), machine_rates)
        operations.append({
            "sno": op.get("sno", len(operations) + 1),
            "process": op.get("process", "Unknown"),
            "machine": op.get("machine", "Unknown"),
            "strokes_per_hr": op.get("strokes_per_hr", 30),
            "rate_per_hr": rate,
        })

    print(f"[Costing] AI suggested {len(operations)} operations")
    return operations


def _rule_based_estimate(
    features: List[Dict[str, Any]],
    metadata: Dict[str, Any],
    machine_rates: Dict[str, float],
) -> List[Dict[str, Any]]:
    """Fallback rule-based operation estimator when AI is unavailable."""
    operations = []
    op_no = 1

    has_od = False
    has_thread_int = False
    has_surface_finish = False
    has_gdt = False
    od_count = 0

    for f in features:
        spec = (f.get("specification") or f.get("spec") or "").lower()
        ftype = (f.get("feature_type") or f.get("type") or "").lower()
        if ftype in ("dimension", "od") or "ø" in spec:
            has_od = True
            od_count += 1
        if "6h" in spec or "6g" in spec:
            has_thread_int = True
        if ftype in ("surface", "surface_finish") or "ra" in spec:
            has_surface_finish = True
        if ftype == "gdt" or "⊙" in spec:
            has_gdt = True

    envelope = metadata.get("part_envelope", {})
    max_od = envelope.get("max_od_mm") or 20.0

    # Parting
    rate = _find_rate("Traub" if max_od <= 25 else "CNC Cutting", machine_rates)
    operations.append({"sno": op_no, "process": "Parting",
                       "machine": "Traub" if max_od <= 25 else "CNC Cutting",
                       "strokes_per_hr": 85, "rate_per_hr": rate})
    op_no += 1

    # CNC
    if has_od:
        rate = _find_rate("CNC", machine_rates)
        operations.append({"sno": op_no, "process": "CNC 1st Setup",
                           "machine": "CNC", "strokes_per_hr": 25, "rate_per_hr": rate})
        op_no += 1
        if od_count > 8:
            operations.append({"sno": op_no, "process": "CNC 2nd Setup",
                               "machine": "CNC", "strokes_per_hr": 18, "rate_per_hr": rate})
            op_no += 1

    # Drilling/Tapping
    if has_thread_int:
        rate = _find_rate("VMC", machine_rates)
        operations.append({"sno": op_no, "process": "Drilling & Tapping",
                           "machine": "VMC/Turret Lathe", "strokes_per_hr": 50, "rate_per_hr": rate})
        op_no += 1

    # Grinding
    if has_surface_finish or has_gdt:
        rate = _find_rate("Centerless Grinding", machine_rates)
        operations.append({"sno": op_no, "process": "Grinding",
                           "machine": "Centerless Grinding", "strokes_per_hr": 50, "rate_per_hr": rate})
        op_no += 1

    return operations


def _find_rate(machine_name: str, rates: Dict[str, float]) -> float:
    """Find machine rate by fuzzy name match."""
    nm = machine_name.upper().strip()
    # Direct match
    for k, v in rates.items():
        if k.upper() == nm:
            return v
    # Partial match
    for k, v in rates.items():
        if k.upper() in nm or nm in k.upper():
            return v
    # Default
    return 200.0


# ---------------------------------------------------------------------------
# 6.  MATERIAL LOOKUP
# ---------------------------------------------------------------------------

def find_material_price(
    material_str: str,
    material_prices: List[Dict[str, Any]],
) -> Optional[Dict[str, Any]]:
    """Find material price entry by grade or alias match."""
    if not material_str:
        return None
    mat_upper = material_str.upper().strip()
    for mp in material_prices:
        if mp["grade"].upper() in mat_upper:
            return mp
        aliases = (mp.get("aliases") or "").split(",")
        for alias in aliases:
            if alias.strip().upper() and alias.strip().upper() in mat_upper:
                return mp
    return None


# ---------------------------------------------------------------------------
# 7.  FULL ESTIMATE (combines everything)
# ---------------------------------------------------------------------------

def generate_full_estimate(
    features: List[Dict[str, Any]],
    metadata: Dict[str, Any],
    machine_rates: Dict[str, float],
    material_prices: List[Dict[str, Any]],
    config: Dict[str, float],
    quantity: int = 3000,
) -> Dict[str, Any]:
    """
    Generate complete cost estimate from feasibility features + metadata.
    Returns all sections ready for Excel generation or API response.
    """
    # Find material
    material_str = metadata.get("material", {}).get("grade", "") if isinstance(metadata.get("material"), dict) else str(metadata.get("material", ""))
    mat_price = find_material_price(material_str, material_prices)
    if not mat_price:
        mat_price = {"grade": material_str, "density_g_cm3": 7.86, "rate_per_kg": 86}

    # Part envelope
    envelope = metadata.get("part_envelope", {})
    max_od = envelope.get("max_od_mm") or 20.0
    total_length = envelope.get("total_length_mm") or 50.0
    stock_allowance = config.get("stock_allowance_mm", 1.0)
    rod_dia = max_od + stock_allowance
    rod_length = total_length + stock_allowance * 2  # allowance on both ends

    # Net weight estimate (from metadata or calculate as 60% of gross)
    density = mat_price.get("density_g_cm3", 7.86)
    gross_est = rod_dia * rod_dia * rod_length * density * 0.786 / 1_000_000
    net_weight = gross_est * 0.6  # rough estimate, user can override

    # Section A
    rm = calculate_raw_material(
        rod_dia_mm=rod_dia,
        rod_length_mm=rod_length,
        net_weight_kg=round(net_weight, 3),
        density_g_cm3=density,
        mat_rate_per_kg=mat_price.get("rate_per_kg", 86),
        zbc=config.get("zbc_default", 30),
        yield_pct=config.get("yield_pct", 85),
    )

    # Section B
    ops = estimate_operations_from_features(features, metadata, machine_rates)
    process = calculate_process_cost(ops, net_weight_kg=net_weight)

    # Section C
    overheads = calculate_overheads(
        rm_cost=rm["part_rm_rate"],
        bop_cost=rm["bop_cost"],
        process_cost=process["total_process_cost"],
        plating_cost=process["plating_cost"],
        annealing_cost=process["annealing_cost"],
        net_weight_kg=net_weight,
        config=config,
    )

    # Total
    total = calculate_total(rm, process, overheads)

    return {
        "part_name": metadata.get("part_name", ""),
        "part_number": metadata.get("drawing_number", ""),
        "material": material_str,
        "material_grade": mat_price.get("grade", ""),
        "quantity": quantity,
        **total,
    }


# ---------------------------------------------------------------------------
# 8.  EXCEL GENERATION
# ---------------------------------------------------------------------------

def generate_cost_sheet_excel(
    costing_data: Dict[str, Any],
    template_path: str,
    output_path: str,
    sheet_name: str = None,
) -> str:
    """
    Generate Cost Break-Up Sheet Excel from template.
    Fills in all calculated values preserving template formatting.
    Returns output file path.
    """
    wb = openpyxl.load_workbook(template_path)

    # Use specified sheet, or find Shaft sheet, or use second sheet (first is often Fork Base)
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    elif "1001540839_Shaft" in wb.sheetnames:
        ws = wb["1001540839_Shaft"]
    elif len(wb.sheetnames) > 1:
        ws = wb[wb.sheetnames[1]]  # second sheet is typically the Shaft template
    else:
        ws = wb.active

    rm = costing_data["raw_material"]
    proc = costing_data["process"]
    oh = costing_data["overheads"]
    tooling = costing_data.get("tooling", {})

    # --- Header ---
    ws["D8"] = costing_data.get("part_name", "")
    ws["D9"] = costing_data.get("part_number", "")
    ws["D10"] = costing_data.get("material", "")
    ws["D12"] = costing_data.get("quantity", 3000)

    # --- Section A: Raw Material ---
    ws["B14"] = rm.get("rod_shape", "ROD")
    ws["C15"] = rm["rod_dia_mm"]
    ws["D15"] = rm["rod_length_mm"]
    # H14 = Gross Weight (formula stays, but we set the inputs)
    ws["H15"] = rm["net_weight_kg"]
    ws["E17"] = rm["zbc"]
    ws["E18"] = rm["yield_pct"]
    ws["H17"] = rm["mat_rate_per_kg"]
    ws["H19"] = rm["bop_cost"]

    # --- Section B: Process Cost ---
    ops = proc["operations"]
    # Template has rows 22-27 for operations. Clear and fill.
    start_row = 22
    max_op_rows = 6  # rows 22-27

    def safe_set(ws, cell_ref, value):
        """Set cell value, skipping merged cells."""
        cell = ws[cell_ref]
        if not isinstance(cell, openpyxl.cell.cell.MergedCell):
            cell.value = value

    # Clear existing operation rows
    for r in range(start_row, start_row + max_op_rows):
        for col in "BCDEFG":
            safe_set(ws, f"{col}{r}", None)

    # Fill operations
    for i, op in enumerate(ops):
        row = start_row + i
        if i >= max_op_rows:
            break
        safe_set(ws, f"B{row}", op["sno"])
        safe_set(ws, f"C{row}", op["process"])
        safe_set(ws, f"D{row}", op["machine"])
        safe_set(ws, f"E{row}", op["strokes_per_hr"])
        safe_set(ws, f"F{row}", op["rate_per_hr"])
        safe_set(ws, f"G{row}", f"=+F{row}/E{row}")

    # Plating — E30 formula =+H15 (net weight), F30 = rate per kg
    safe_set(ws, "E30", "=+H15")
    safe_set(ws, "F30", proc.get("plating_rate_per_kg", 0))
    # Annealing
    safe_set(ws, "E32", proc.get("surface_area_sqin", 0))
    safe_set(ws, "F32", proc.get("annealing_rate_per_sqin", 0))

    # --- Section C: Overheads ---
    safe_set(ws, "F36", oh["rejection_pct"])
    safe_set(ws, "F37", oh["icc_pct"])
    safe_set(ws, "F38", oh["overheads_pct"])
    safe_set(ws, "F39", oh["profit_pct"])
    safe_set(ws, "F40", oh["packing_pct"])
    safe_set(ws, "F41", oh["freight_per_kg"])
    safe_set(ws, "F42", oh["inspection_pct"])

    # --- Section D: Tooling ---
    safe_set(ws, "E47", tooling.get("hydraulic_fixture", 0))
    safe_set(ws, "E48", tooling.get("gauges", 0))
    safe_set(ws, "E49", tooling.get("investment_casting_tool", 0))

    # Remove extra sheets (keep only the one we filled)
    for sn in wb.sheetnames:
        if sn != ws.title:
            del wb[sn]

    # Rename sheet to part number
    part_no = costing_data.get("part_number", "Part")
    ws.title = f"{part_no}_{costing_data.get('part_name', '')}"[:31]

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return output_path
